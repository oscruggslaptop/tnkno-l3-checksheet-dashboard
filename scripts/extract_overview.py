import json
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


PALETTE = {
    "slate": "#1F2937",
    "amber": "#FFC000",
    "amberDeep": "#FFCA08",
    "red": "#FF0000",
    "green": "#92D050",
    "grey": "#D9D9D9",
    "white": "#FFFFFF",
}

FAILURE_SHEETS = [
    {
        "sheet": "Module",
        "status_col": 6,
        "source": "Conveyors/Panels/Field Devices",
        "fields": ["category", "location", "item", "description", "type", None, None, None, "detail"],
    },
    {
        "sheet": "System Logic",
        "status_col": 7,
        "source": "System Logic",
        "fields": ["id", "category", "test", "location", "item", "description", None, "detail"],
    },
    {
        "sheet": "Stats",
        "status_col": 6,
        "source": "HMI Statistics",
        "fields": ["category", "subCategory", "item", "description", "type", None, "detail"],
    },
    {
        "sheet": "Stats (2)",
        "status_col": 6,
        "source": "HMI Statistics",
        "fields": ["category", "subCategory", "item", "description", "type", None, "detail"],
    },
    {
        "sheet": "Reporting Summary",
        "status_col": 7,
        "source": "BAU Push / Reporting",
        "fields": ["id", "category", "subCategory", "item", "type", "description", None, "detail"],
    },
    {
        "sheet": "HSLA",
        "status_col": 7,
        "source": "HSLA",
        "fields": ["id", "category", "step", "item", "description", "action", None, "detail"],
    },
]


def value(ws, cell):
    raw = ws[cell].value
    if isinstance(raw, datetime):
        return raw.isoformat()
    return raw


def pct(raw):
    return raw if isinstance(raw, (int, float)) else None


def metric(label, complete, total=None, passed=None, failed=None, tone="neutral"):
    return {
        "label": label.replace("\xa0", " ").strip(),
        "complete": complete,
        "total": total,
        "passed": passed,
        "failed": failed,
        "tone": tone,
    }


def clean(raw):
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return raw.isoformat()
    text = str(raw).replace("\xa0", " ").strip()
    return text if text else None


def first_meaningful(*values):
    for item in values:
        if item is not None and str(item).strip():
            return item
    return None


def extract_failures(wb):
    failures = []
    for definition in FAILURE_SHEETS:
        if definition["sheet"] not in wb.sheetnames:
            continue
        ws = wb[definition["sheet"]]
        for row_number, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row_number == 1:
                continue
            status = clean(row[definition["status_col"] - 1] if len(row) >= definition["status_col"] else None)
            if str(status).lower() not in {"f", "fail", "failed"}:
                continue
            item = {}
            for index, field in enumerate(definition["fields"]):
                if field and index < len(row):
                    item[field] = clean(row[index])
            failures.append(
                {
                    "source": definition["source"],
                    "sheet": definition["sheet"],
                    "row": row_number,
                    "category": first_meaningful(item.get("category"), item.get("subCategory"), item.get("test")),
                    "location": first_meaningful(item.get("location"), item.get("step")),
                    "item": first_meaningful(
                        item.get("item"),
                        item.get("description"),
                        item.get("metric"),
                        item.get("test"),
                        f"Row {row_number}",
                    ),
                    "description": first_meaningful(item.get("description"), item.get("action"), item.get("type")),
                    "detail": first_meaningful(item.get("detail"), item.get("action")),
                    "status": "Fail",
                }
            )
    return failures


def main():
    if len(sys.argv) != 2:
        raise SystemExit("usage: extract_overview.py <workbook.xlsm>")

    workbook_path = Path(sys.argv[1])
    wb = load_workbook(workbook_path, read_only=True, keep_vba=True, data_only=True)
    if "Overview" not in wb.sheetnames:
        raise SystemExit(f"Overview sheet not found. Sheets: {', '.join(wb.sheetnames)}")
    ws = wb["Overview"]

    metrics = [
        metric(value(ws, "B3"), pct(value(ws, "C3")), tone="primary"),
        metric(value(ws, "B4"), pct(value(ws, "C4")), passed=pct(value(ws, "D4")), failed=pct(value(ws, "E4"))),
        metric(value(ws, "B5"), pct(value(ws, "C5")), passed=pct(value(ws, "D5")), failed=pct(value(ws, "E5"))),
        metric(value(ws, "B6"), pct(value(ws, "C6")), tone="primary"),
        metric(value(ws, "B7"), pct(value(ws, "C7")), passed=pct(value(ws, "D7")), failed=pct(value(ws, "E7"))),
        metric(value(ws, "B8"), pct(value(ws, "C8")), passed=pct(value(ws, "D8")), failed=pct(value(ws, "E8"))),
        metric(value(ws, "B9"), pct(value(ws, "C9")), passed=pct(value(ws, "D9")), failed=pct(value(ws, "E9"))),
        metric(value(ws, "B10"), pct(value(ws, "C10")), tone="primary"),
        metric(value(ws, "B11"), None, value(ws, "C11"), value(ws, "D11"), value(ws, "E11"), tone="readiness"),
        metric(value(ws, "B12"), pct(value(ws, "C12")), tone="primary"),
        metric(value(ws, "B15"), pct(value(ws, "C15")), tone="pass"),
    ]

    breakdown = []
    for row in range(4, 7):
        breakdown.append(
            {
                "type": value(ws, f"G{row}"),
                "total": value(ws, f"H{row}"),
                "remaining": value(ws, f"I{row}"),
                "complete": pct(value(ws, f"J{row}")),
            }
        )

    moduleRows = []
    for row in range(19, 32):
        label = value(ws, f"B{row}")
        if label:
            moduleRows.append(
                {
                    "label": str(label).replace("\xa0", " ").strip(),
                    "total": value(ws, f"C{row}"),
                    "passed": value(ws, f"D{row}"),
                    "failed": value(ws, f"E{row}"),
                }
            )

    payload = {
        "title": "L3 Checksheet",
        "module": value(ws, "B2"),
        "sheet": "Overview",
        "sourceFile": workbook_path.name,
        "lastModified": datetime.fromtimestamp(workbook_path.stat().st_mtime).isoformat(),
        "palette": PALETTE,
        "summary": {
            "overallComplete": pct(value(ws, "C3")),
            "overallPass": pct(value(ws, "C15")),
            "plannedStart": value(ws, "C13"),
            "actualStart": value(ws, "C14"),
        },
        "metrics": metrics,
        "breakdown": breakdown,
        "moduleRows": moduleRows,
        "failures": extract_failures(wb),
    }
    print(json.dumps(payload, default=str))


if __name__ == "__main__":
    main()
